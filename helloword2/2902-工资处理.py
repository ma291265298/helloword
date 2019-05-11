import openpyxl

wb = openpyxl.load_workbook('工资统计.xlsx')
sheet1 = wb['sheet1']
sheet2 = wb['实验教学任务']
sheet3 = wb['实验工作量']
sheet4 = wb['理论教学任务']
sheet5 = wb['理论工作量']
sheet6 = wb['工作量汇总']
sheet1.title = '教工基本信息'

theory = {}
experiment = {}

for i in range(2, sheet2.max_row + 1):
    course_name = sheet2.cell(i, 1).value
    time = sheet2.cell(i, 2).value
    class_number = sheet2.cell(i, 3).value
    class_name = sheet2.cell(i, 4).value
    teacher_id = sheet2.cell(i, 5).value
    teacher_name = sheet2.cell(i, 6).value
    sheet3.cell(i, 1).value = teacher_id
    sheet3.cell(i, 2).value = teacher_name
    sheet3.cell(i, 3).value = course_name
    sheet3.cell(i, 4).value = class_name
    sheet3.cell(i, 5).value = time
    sheet3.cell(i, 6).value = class_number
    coefficient = 0.7 * (1 + (class_number - 40) * 0.015)
    # sheet3.cell(i,7).value=coefficient
    coefficients = '=0.7 * (1 + (F' + str(i) + ' - 40) * 0.015)'
    sheet3.cell(i, 7).value = coefficients
    workload = int(coefficient * time * 100 + 0.5) / 100
    sheet3.cell(i, 8).value = workload

    try:
        experiment[str(teacher_id)] = experiment[str(teacher_id)] + float(workload)
    except:
        experiment[str(teacher_id)] = float(workload)

for i in range(2, sheet4.max_row + 1):
    dic = {'是': 0.9, '否': 1}
    course_name = sheet4.cell(i, 1).value
    time = sheet4.cell(i, 2).value
    class_number = sheet4.cell(i, 3).value
    class_name = sheet4.cell(i, 4).value
    teacher_id = sheet4.cell(i, 5).value
    teacher_name = sheet4.cell(i, 6).value
    class_quality = sheet4.cell(i, 7).value
    sheet5.cell(i, 1).value = teacher_id
    sheet5.cell(i, 2).value = teacher_name
    sheet5.cell(i, 3).value = course_name
    sheet5.cell(i, 4).value = time
    sheet5.cell(i, 5).value = class_quality
    sheet5.cell(i, 6).value = dic[class_quality]
    sheet5.cell(i, 7).value = class_number
    if class_number <= 50:
        number_quality = 1
    else:
        number_quality = 1 + (class_number - 50) / 150
    sheet5.cell(i, 8).value = number_quality
    sheet5.cell(i, 9).value = class_name
    workload = int((float(time) * dic[class_quality] * float(number_quality)) * 100 + 0.5) / 100
    sheet5.cell(i, 10).value = workload

    try:
        theory[str(teacher_id)] = theory[str(teacher_id)] + float(workload)
    except:
        theory[str(teacher_id)] = float(workload)

for i in range(2, sheet1.max_row + 1):
    teacher_id = sheet1.cell(i, 1).value
    teacher_name = sheet1.cell(i, 2).value
    try:
        sumexperiment = experiment[teacher_id]
    except:
        sumexperiment=0
    try:
        sumtheory = theory[teacher_id]
    except:
        sumtheory=0
    sumworkload = sumexperiment + sumtheory
    if sumworkload <= 100:
        money = 50 * sumworkload
    else:
        money = 5000 + 70 * (sumworkload - 100)
    sheet6.cell(i, 1).value = teacher_id
    sheet6.cell(i, 2).value = teacher_name
    sheet6.cell(i, 3).value = sumexperiment
    sheet6.cell(i, 4).value = sumtheory
    sheet6.cell(i, 5).value = sumworkload
    sheet6.cell(i, 6).value = money

wb.save('2902工资处理.xlsx')
