import openpyxl


# 作业成绩
def task():
    dic = {'A': 95, 'B': 85, 'C': 75, 'D': 65, 'E': 0}
    wb = openpyxl.load_workbook('作业成绩.xlsx')
    sheet1 = wb['平时作业 ']

    for i in range(5, sheet1.max_row - 1):
        grade = 0
        for j in range(4, 12):
            grades = str(sheet1.cell(i, j).value)
            grade += dic[grades]
        avegrade = float(grade / 8)
        sheet1.cell(i, 12).value = avegrade
    wb.save('2901作业成绩.xlsx')


# 实验成绩
def experiment():
    dic = {'A': 95, 'B': 85, 'C': 75, 'D': 65, 'E': 0}
    wb = openpyxl.load_workbook('实验成绩.xlsx')
    sheet1 = wb['实验报告']
    sheet2 = wb['实验成绩']
    for i in range(6, sheet1.max_row - 1):
        grade = 0
        for j in range(4, 12):
            grades = str(sheet1.cell(i, j).value)
            grade += dic[grades]
        avegrade = float(grade / 8)
        sheet1.cell(i, 12).value = avegrade
        sheet2.cell(i, 6).value = avegrade
        sum = 0.1 * float(sheet2.cell(i, 4).value) + 0.1 * float(sheet2.cell(i, 5).value) + 0.8 * float(
            sheet2.cell(i, 6).value)
        sheet2.cell(i, 7).value = sum
    wb.save('2901实验成绩.xlsx')


# 成绩汇总
def Summary():
    wb1 = openpyxl.load_workbook('2901作业成绩.xlsx')
    sheet1 = wb1['平时作业 ']
    wb2 = openpyxl.load_workbook('2901实验成绩.xlsx')
    sheet2 = wb2['实验成绩']
    wb3 = openpyxl.load_workbook('成绩汇总.xlsx')
    sheet3 = wb3['Sheet1']
    for i in range(5, 32):
        tasks = float(sheet1.cell(i, 12).value)
        experiments = float(sheet2.cell(i + 1, 7).value)
        sheet3.cell(i + 1, 5).value = tasks
        sheet3.cell(i + 1, 8).value = experiments
        midterm = float(sheet3.cell(i + 1, 6).value)
        terminal = float(sheet3.cell(i + 1, 7).value)
        sumgrade = 0.1 * tasks + 0.1 * midterm + 0.6 * terminal + 0.2 * experiments
        sheet3.cell(i + 1, 9).value = sumgrade
    wb3.save('2901成绩汇总.xlsx')

task()
experiment()
Summary()