import openpyxl
from openpyxl.chart import BarChart, Series, Reference

wb = openpyxl.load_workbook('工资处理.xlsx')
sheet = wb['工作量汇总']
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

sheet2.cell(1, 1).value = '教师姓名'
sheet2.cell(1, 2).value = '工作量工资'

for i in range(2, 14):
    sheet2.cell(i, 1).value = sheet.cell(i, 2).value
    sheet2.cell(i, 2).value = sheet.cell(i, 6).value

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "工资图表"
chart1.y_axis.title = '工资'
chart1.x_axis.title = '姓名'

data = Reference(sheet2, min_col=2, min_row=1, max_row=13, max_col=2)
cats = Reference(sheet2, min_col=1, min_row=2, max_row=13)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
sheet2.add_chart(chart1, "D2")

wb2.save('2903生成图表.xlsx')
