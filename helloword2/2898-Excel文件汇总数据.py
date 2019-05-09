import openpyxl

wb = openpyxl.load_workbook('每个人的爱好.xlsx')
sheet = wb['Sheet1']

lastCol = sheet.max_row + 2
sheet.cell(row=1, column=lastCol).value = '所有爱好'
for i in range(2, sheet.max_row + 1):
    lst = []

    for j in range(2, sheet.max_column + 1):
        value = sheet.cell(row=i, column=j).value
        if value == '是':
            x = sheet.cell(row=1, column=j).value
            lst.append(x)
    sheet.cell(row=i, column=lastCol).value = '，'.join(lst)
wb.save('2898.xlsx')
