import openpyxl
import random

wb = openpyxl.Workbook()
sheet = wb.active

name1 = ['赵', '钱', '孙', '李']
name2 = ['伟', '昀', '琛', '东']
name3 = ['', '坤', '艳', '志']

course = ['语文', '数学', '英语']

sheet.cell(row=1, column=1).value = '姓名'
sheet.cell(row=1, column=2).value = '课程'
sheet.cell(1, 3).value = '成绩'
dic = {}
for i in range(2, 100):
    name = name1[random.randint(0, 3)] + name2[random.randint(0, 3)] + name3[random.randint(0, 3)]
    sheet.cell(i, 1).value = name
    cour = course[random.randint(0, 2)]
    sheet.cell(i, 2).value = cour
    grade = random.randint(0, 100)
    sheet.cell(i, 3).value = grade
    naco = name + '，' + cour
    try:
        if dic[naco] < grade:
            dic[naco] = grade
    except:
        dic[naco] = grade
wb.save('test.xlsx')

wb2 = openpyxl.Workbook()
sheet2 = wb2.active

lis = list(dic.keys())
lis2 = list(dic.values())
sumlist = []
for i in range(len(lis)):
    sumlist.append(str(lis[i]) + '，' + str(lis2[i]))
sumlist = sorted(sumlist)
sheet2.cell(row=1, column=1).value = '姓名'
sheet2.cell(row=1, column=2).value = '课程'
sheet2.cell(1, 3).value = '最高成绩'
for i in range(len(lis)):
    name, cour, grade = str(sumlist[i]).split('，')
    sheet2.cell(i + 2, 1).value = name
    sheet2.cell(i + 2, 2).value = cour
    sheet2.cell(i + 2, 3).value = grade
wb2.save('result.xlsx')
