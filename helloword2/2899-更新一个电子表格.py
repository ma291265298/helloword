import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

for i in range(2, sheet.max_row + 1):
    value = sheet.cell(row=i, column=1).value
    if value == 'Lemon':
        sheet.cell(row=i, column=2).value = 1.27
    elif value == 'Garlic':
        sheet.cell(row=i, column=2).value = 3.07
    elif value == 'Celery':
        sheet.cell(row=i, column=2).value = 1.19
wb.save('2899.xlsx')
